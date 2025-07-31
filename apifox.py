# -*- coding: UTF-8 -*-
import os
from datetime import datetime
import subprocess
import requests
import json  # 标准库 json 主要用于 JSON 数据的读取和写入，而不提供直接的 JSONPath 功能
from jsonpath_ng import jsonpath, parse  # 专门的 JSONPath 解析库
import configparser
import openpyxl
import re
import mysql_operation
from typing import Dict, List, Optional, Tuple
from dataclasses import dataclass
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
import time
"""
版本更新说明：
V1.0
基础框架实现

V1.1
apifox对apifox-report.json执行日志文件内容格式做了调整，进行适配

V1.2
对输出的执行失败结果做一个归类拆分，先输出断言预期错误的内容，再输出接口耗时超时的内容

V1.3
对每次执行的内容写入数据表apifox_fail_case做落盘，到每周、月结束做异常次数统计后输出

V1.4
1.兼容新版本的apifox CLI，可以使用新的CICD命令运行测试用例集
2.兼容新版本CLI执行后的json日志的格式调整，调整参数取值

V2.0
整体代码结构调整，重构所有的函数

V3.0
1.支持多线程运行
2.支持指定对应项目的接口用例执行
"""

@dataclass
class ApifoxConfig:
    """Apifox配置类"""
    access_token: str = "APS-STHxxxxxxxxxxxxxxxx"
    excel_path: str = 'apifox_url.xlsx'
    cli_path: str = "D:/Nodejs/node.exe C:/Users/AppData/Roaming/npm/node_modules/apifox-cli/bin/cli.js"
    # Webhook配置
    webhook_feishu_url_test: str = "https://open.feishu.cn/open-apis/bot/v2/hook/xxxx"
    webhook_feishu_url_online: str = "https://open.feishu.cn/open-apis/bot/v2/hook/xxxxxx"
    webhook_wechat_url_online: str = "https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key=xxxxx"
    # 线程池配置
    max_workers: int = 2  # 最大线程数，线上和线下各一个线程

    def get_webhook_url(self, online: bool, type: str = 'feishu') -> str:
        """获取webhook URL
        
        Args:
            online: 是否为线上环境
            type: 消息类型，'feishu'或'wechat'
            
        Returns:
            str: webhook URL
        """
        if online:
            return self.webhook_feishu_url_online if type == 'feishu' else self.webhook_wechat_url_online
        return self.webhook_feishu_url_test

    def get_message_data(self, message: str, type: str = 'feishu') -> dict:
        """获取消息数据格式
        
        Args:
            message: 消息内容
            type: 消息类型，'feishu'或'wechat'
            
        Returns:
            dict: 消息数据
        """
        if type == 'feishu':
            return {
                "msg_type": "text",
                "content": {"text": message}
            }
        return {
            "msgtype": "text",
            "text": {"content": message}
        }

class apifox_auto_test():
    def __init__(self, config: Optional[ApifoxConfig] = None):
        """初始化Apifox自动化测试类
        
        Args:
            config: Apifox配置对象，如果为None则使用默认配置
        """
        # 初始化计数器
        self.total_case: int = 0
        self.total_fail_case: int = 0
        self.total_online_fail_case: int = 0
        self.jsonfile_list: List[str] = []
        self.total_fail_case_info: Dict = {}
        
        # 线程安全的锁
        self._lock = threading.Lock()
        
        # 加载配置
        self.config = config or ApifoxConfig()
        
        # 加载测试用例数据
        try:
            self.online_cases, self.offline_cases = self._load_test_cases()
        except Exception as e:
            print(f"加载测试用例失败: {str(e)}")
            self.online_cases = {}
            self.offline_cases = {}

    def _load_test_cases(self) -> Tuple[Dict[str, str], Dict[str, str]]:
        """从Excel文件加载测试用例，分离线上和线下用例
        
        Returns:
            Tuple[Dict[str, str], Dict[str, str]]: (线上用例字典, 线下用例字典)
        """
        online_cases = {}
        offline_cases = {}
        
        try:
            workbook = openpyxl.load_workbook(self.config.excel_path)
            worksheet = workbook.active
            max_row = worksheet.max_row

            for row in range(1, max_row + 1):
                key_cell = worksheet.cell(row=row, column=1)
                value_cell = worksheet.cell(row=row, column=8)
                
                if key_cell.value is None:
                    continue
                    
                command = self._parse_command(value_cell.value)
                if command:
                    # 根据用例名称判断是否为线上用例
                    if "(线上)" in key_cell.value or "（线上）" in key_cell.value:
                        online_cases[key_cell.value] = command
                    else:
                        offline_cases[key_cell.value] = command
                    
            workbook.close()
            return online_cases, offline_cases
            
        except Exception as e:
            print(f"读取Excel文件失败: {str(e)}")
            return {}, {}

    def _parse_command(self, command_value: Optional[str]) -> Optional[str]:
        """解析命令字符串
        
        Args:
            command_value: 命令字符串
            
        Returns:
            Optional[str]: 解析后的命令，如果解析失败则返回None
        """
        if not command_value:
            return None
            
        try:
            if command_value.startswith("apifox run --access-token"):
                # 新版命令格式处理
                parts = command_value.split(' ')
                access_token_index = parts.index('--access-token')
                r_index = parts.index('-r')
                return ' '.join(parts[access_token_index:r_index])
            else:
                # 旧版命令格式处理
                parts = command_value.split(' ')
                return parts[2] if len(parts) >= 3 else None
                
        except (ValueError, IndexError):
            return None

    def run_command(self,
                    command="https://api.apifox.cn/api/v1/projects/2875535/api-test/ci-config/375963/detail?token=x4"):
        """执行apifox CLI的命令,
        新版的命令格式：
        apifox run --access-token APS-STHxxxxxxxxxxxxxx -t 555555 -e 12222222 -n 1 -r html,cli
        apifox run --access-token $APIFOX_ACCESS_TOKEN -t 566666 -e 12222222 -n 1 -r html,cli
        旧版的命令格式：
        apifox run https://api.apifox.com/api/v1/projects/2875419/api-test/ci-config/388754/detail?token=xxxxxxxxxx -r html,cli
        使用 Apifox 的 Access Token 运行指定的测试场景或测试场景目录，示例：
        apifox run --access-token $APIFOX_ACCESS_TOKEN -t 637132 -e 358171 -d 3497013 -r html,cli --database-connection ./database-connections.json"""
        now = datetime.now()
        date_time = now.strftime("%Y-%m-%d_%H-%M-%S")
        filename = "apifox-report-" + f"{date_time}"
        
        if "$APIFOX_ACCESS_TOKEN" in command:
            command = command.replace("$APIFOX_ACCESS_TOKEN", self.config.access_token)
        apifox_command = self.config.cli_path + " run " + command + " -r json" + " --out-file {}".format(filename) + " --upload-report" + " --verbose"
        
        # 输出到脚本目录下\apifox-reports文件夹
        try:
            result = subprocess.check_output(apifox_command, shell=True, stderr=subprocess.STDOUT,
                                             universal_newlines=False)
            #  将subprocess.check_output中的universal_newlines=True参数更改为False，这将返回未解码的字节字符串，而不是尝试将其解码为文本
            print("{}:命令执行成功:".format(date_time))
            print(result.decode("utf-8"))
            
            # 使用线程锁保护共享资源
            with self._lock:
                self.jsonfile_list.append(filename)
                
        except subprocess.CalledProcessError as e:
            print("{}:命令执行完成:".format(date_time))
            print(e.output.decode("utf-8"))
            
            # 使用线程锁保护共享资源
            with self._lock:
                self.jsonfile_list.append(filename)
                
        except Exception as e:
            print("{}:发生错误:".format(date_time))
            print(str(e))

    def deal_with_fail_reason(self, fail_reason):
        """针对返回的结果，做一层处理后输出
        $.meta.total 小于 40110 | AssertionError: expected 40110 to be below 40110
        接口响应小于1000ms | AssertionError: expected 1167 to be below 1000
        list内容不为[] | AssertionError: expected [] to not deeply equal []
        list内容不为空 | AssertionError: expected '' not to be empty
        """
        pattern = re.compile(r'expected (\d+) to be below (\d+)')
        # 查找匹配项
        match = pattern.search(fail_reason)
        if match:
            first_number = int(match.group(1))
            second_number = int(match.group(2))
            first_number = first_number / 1000
            first_number = round(first_number, 2)
            second_number = second_number / 1000
            second_number = round(second_number, 2)
            fail_reason = "接口执行耗时{}秒，标准需要小于{}秒,要优化.".format(first_number, second_number)
            return fail_reason
        pattern = re.compile(r'expected (.+) to deeply equal (.+)')
        # 查找匹配项
        match = pattern.search(fail_reason)
        if match:
            first_number = match.group(1)
            second_number = match.group(2)
            fail_reason = "接口返回预期：{}，实际：{},存在偏差请检查.".format(second_number, first_number)
            return fail_reason
        pattern = re.compile(r'expected (.+) to not deeply equal (.+)')
        # 查找匹配项
        match = pattern.search(fail_reason)
        if match:
            first_number = match.group(1)
            second_number = match.group(2)
            fail_reason = "接口返回预期为非：{}，实际：{},存在偏差请检查.".format(second_number, first_number)
            return fail_reason
        return fail_reason

    def json_analyse(self, filename="apifox-report-2023-09-12-17-20-08-602-0.json"):
        """分析输出的json报告"""
        path = "apifox-reports/"
        is_online_case = 0
        date_part = filename.split("_")[0].split("-")[-3:]
        case_occurrence_time = "-".join(date_part)
        file_path = path + filename
        if ".json" not in file_path:
            file_path += ".json"
        if os.path.exists(file_path):
            try:
                with open(file_path, 'r', encoding='utf-8') as json_file:
                    # 使用 json.load() 解析 JSON 文件内容为 Python 数据结构
                    data = json.load(json_file)
                # 现在，'data' 变量包含了 JSON 文件中的数据，可以像访问字典一样访问其中的内容
                # total_count = data['result']['stats']['requests']['total']
                # fail_count = data['result']['stats']['requests']['failed']
                total_count = data['result']['stats']['steps']['total']
                fail_count = data['result']['stats']['steps']['failed']
                result_dict_error = {}
                result_dict_timeout = {}
                jsonpath_expr = parse("$.collection.name")  # 取外部的整个测试用例集的名字
                # 使用 JSONPath 表达式提取数据
                matches_fail_case_parent = [match.value if match.value else 'None' for match in
                                            jsonpath_expr.find(data)]
                if matches_fail_case_parent:
                    matches_fail_case_parent = matches_fail_case_parent[0]
                if "(线上)" in matches_fail_case_parent or "（线上）" in matches_fail_case_parent:
                    is_online_case = 1
                # 定义 JSONPath 表达式
                if fail_count > 0:
                    # jsonpath_expr = parse("$.result.steps[*].id")
                    # jsonpath_expr = parse("$.collection.item[0].item[*].id")
                    jsonpath_expr = parse("$.run.failures[*].cursor.ref")
                    case_id = [match.value for match in jsonpath_expr.find(data)]
                    # jsonpath_expr = parse("$.result.steps[*].name")
                    # jsonpath_expr = parse("$.collection.item[0].item[*].name")
                    jsonpath_expr = parse("$.run.failures[*].source.name")
                    case_name = [match.value for match in jsonpath_expr.find(data)]
                    # jsonpath_expr = parse("$.result.steps[*].metaInfo.httpApiPath")
                    # jsonpath_expr = parse("$.collection.item[0].item[*].metaInfo.httpApiPath")
                    jsonpath_expr = parse("$.run.failures[*].source.scope.httpApiPath")
                    case_url = [match.value for match in jsonpath_expr.find(data)]
                    jsonpath_expr = parse("$.run.failures[*].error.message")  # 错误信息
                    matches_fail_reason = [match.value for match in jsonpath_expr.find(data)]
                    jsonpath_expr = parse("$.run.failures[*].error")  # 错误判定备注
                    matches_fail_comment = [match.value.get('test', '无断言备注') for match in jsonpath_expr.find(data)]
                    
                    # 直接组合失败用例信息
                    matches_fail_case = []
                    for i in range(len(case_id)):
                        matches_fail_case.append({
                            'occurrence_time': case_occurrence_time,
                            'case_name': case_name[i],
                            'case_url': case_url[i],
                            "fail_comment": matches_fail_comment[i],
                            'fail_reason': matches_fail_reason[i],
                            'is_online': is_online_case
                        })

                    for fail_case in matches_fail_case:
                        fail_case_name = fail_case['case_name']
                        fail_comment = fail_case['fail_comment']
                        fail_reason = fail_case['fail_reason']
                        fail_reason = self.deal_with_fail_reason(fail_reason)
                        fail_path = fail_case['case_url']
                        is_online = fail_case['is_online']
                        
                        if "执行耗时" not in fail_reason:  # 如果不是耗时错误，拼在第一个字典中
                            result_dict_error[fail_case_name] = {
                                "断言内容": fail_comment,
                                "错误内容": fail_reason,
                                "测试用例集": matches_fail_case_parent,
                                "接口地址": fail_path,
                                "执行时间": case_occurrence_time,
                                "是否线上": is_online
                            }
                        else:
                            result_dict_timeout[fail_case_name] = {
                                "断言内容": fail_comment,
                                "错误内容": fail_reason,
                                "测试用例集": matches_fail_case_parent,
                                "接口地址": fail_path,
                                "执行时间": case_occurrence_time,
                                "是否线上": is_online
                            }
                result_dict = {**result_dict_error, **result_dict_timeout}
                # result_dict = result_dict_error
                return total_count, fail_count, result_dict, is_online_case
            except json.decoder.JSONDecodeError as e:
                print(f"JSON解析错误：{str(e)}")
                return False
            except Exception as e:
                print(e)
                return False

    def send_message(self, message: str = "", online: bool = False, type: str = 'feishu') -> None:
        """通过webhook发送消息
        
        Args:
            message: 消息内容
            online: 是否为线上环境
            type: 消息类型，'feishu'或'wechat'
        """
        try:
            # 获取webhook URL
            webhook_url = self.config.get_webhook_url(online, type)
            
            # 获取消息数据
            message_data = self.config.get_message_data(message, type)
            
            # 发送请求
            headers = {'Content-Type': 'application/json'} if type == 'feishu' else None
            response = requests.post(
                webhook_url,
                json=message_data if type == 'wechat' else None,
                data=json.dumps(message_data) if type == 'feishu' else None,
                headers=headers
            )
            
            # 检查响应结果
            if response.status_code == 200:
                print("消息发送成功")
            else:
                print(f"消息发送失败. 状态码: {response.status_code}, 响应: {response.text}")
                
        except Exception as e:
            print(f"发送消息时发生错误: {str(e)}")

    def _execute_cases_in_thread(self, cases: Dict[str, str], thread_name: str) -> None:
        """在线程中执行测试用例
        
        Args:
            cases: 要执行的用例字典
            thread_name: 线程名称
        """
        print(f"{thread_name} 开始执行，共 {len(cases)} 个用例")
        start_time = time.time()
        
        for case_name, command in cases.items():
            print(f"{thread_name} 执行用例: {case_name}")
            self.run_command(command)
            
        end_time = time.time()
        print(f"{thread_name} 执行完成，耗时: {end_time - start_time:.2f}秒")

    def _execute_test_cases_parallel(self, run_online_case_only: bool) -> None:
        """并行执行测试用例
        
        Args:
            run_online_case_only: 是否只执行线上用例
        """
        if run_online_case_only:
            # 只执行线上用例
            if self.online_cases:
                self._execute_cases_in_thread(self.online_cases, "线上用例线程")
            return
            
        # 使用线程池并行执行线上和线下用例
        with ThreadPoolExecutor(max_workers=self.config.max_workers) as executor:
            futures = []
            
            # 提交线上用例执行任务
            if self.online_cases:
                future_online = executor.submit(
                    self._execute_cases_in_thread, 
                    self.online_cases, 
                    "线上用例线程"
                )
                futures.append(future_online)
            
            # 提交线下用例执行任务
            if self.offline_cases:
                future_offline = executor.submit(
                    self._execute_cases_in_thread, 
                    self.offline_cases, 
                    "线下用例线程"
                )
                futures.append(future_offline)
            
            # 等待所有任务完成
            for future in as_completed(futures):
                try:
                    future.result()
                except Exception as e:
                    print(f"线程执行出错: {str(e)}")

    def _execute_test_cases(self, run_online_case_only: bool) -> None:
        """执行测试用例（兼容旧版本）
        
        Args:
            run_online_case_only: 是否只执行线上用例
        """
        # 使用新的并行执行方法
        self._execute_test_cases_parallel(run_online_case_only)

    def _process_test_results(self) -> None:
        """处理测试结果"""
        for file in self.jsonfile_list:
            if not file:
                continue
                
            result = self.json_analyse(file)
            if not result:
                continue
                
            total_count, fail_count, result_dict, is_online_case = result
            
            # 使用线程锁保护共享计数器
            with self._lock:
                self.total_case += total_count
                self.total_fail_case += fail_count
                self.total_fail_case_info.update(result_dict)
                if is_online_case:
                    self.total_online_fail_case += fail_count

    def _generate_summary_message(self, run_online_case_only: bool) -> str:
        """生成测试总结消息
        
        Args:
            run_online_case_only: 是否只执行线上用例
            
        Returns:
            str: 总结消息
        """
        message = "共测试接口用例{}条，失败{}条，其中线上{}条".format(
            self.total_case, self.total_fail_case, self.total_online_fail_case
        )
        
        if run_online_case_only:
            message = "本次执行只运行线上用例，" + message
            
        if self.total_fail_case == 0:
            return message + "，震惊，再接再厉！"
            
        if self.total_online_fail_case == 0:
            return message + "，线上没有出问题也不错！再接再厉！"
            
        return message + "，失败的线上用例如下:\n"

    def _generate_fail_case_messages(self) -> Tuple[str, str]:
        """生成失败用例消息
        
        Returns:
            Tuple[str, str]: (线上失败用例消息, 线下失败用例消息)
        """
        online_message = ""
        offline_message = "共测试接口用例{}条，失败{}条，失败的线下用例如下:\n".format(
            self.total_case, self.total_fail_case
        )
        
        if self.total_fail_case == 0:
            return online_message, offline_message
            
        online_index = 1
        offline_index = 1
        
        for key, value in self.total_fail_case_info.items():
            case_message = "{}.{}: {}\n".format(online_index if "(线上)" in value['测试用例集'] or "（线上）" in value['测试用例集'] else offline_index, key, value)
            
            if "(线上)" in value['测试用例集'] or "（线上）" in value['测试用例集']:
                online_message += case_message
                online_index += 1
            else:
                offline_message += case_message
                offline_index += 1
                
        return online_message, offline_message

    def _get_case_statistics(self) -> str:
        """获取用例统计信息
        
        Returns:
            str: 统计信息字符串
        """
        online_count = len(self.online_cases)
        offline_count = len(self.offline_cases)
        total_count = online_count + offline_count
        
        return f"总用例数: {total_count} (线上: {online_count}, 线下: {offline_count})"

    def _filter_cases_by_project(self, cases: Dict[str, str], project_keywords: List[str]) -> Dict[str, str]:
        """根据项目关键词过滤用例
        
        Args:
            cases: 用例字典
            project_keywords: 项目关键词列表
            
        Returns:
            Dict[str, str]: 过滤后的用例字典
        """
        if not project_keywords:
            return cases
            
        filtered_cases = {}
        for case_name, command in cases.items():
            if any(keyword.lower() in case_name.lower() for keyword in project_keywords):
                filtered_cases[case_name] = command
                
        return filtered_cases

    def total_test(self, send_online_message: bool = False, run_online_case_only: bool = False, 
                   project_keywords: Optional[List[str]] = None) -> None:
        """执行所有测试用例并生成报告
        
        Args:
            send_online_message: 是否发送线上消息
            run_online_case_only: 是否只执行线上用例
            project_keywords: 项目关键词列表，用于过滤特定项目的用例
        """
        # 根据项目关键词过滤用例
        if project_keywords:
            self.online_cases = self._filter_cases_by_project(self.online_cases, project_keywords)
            self.offline_cases = self._filter_cases_by_project(self.offline_cases, project_keywords)
            print(f"根据项目关键词 {project_keywords} 过滤用例")
        
        # 显示用例统计信息
        print(f"用例统计: {self._get_case_statistics()}")
        
        # 重置计数器
        self.jsonfile_list = []
        self.total_fail_case_info = {}
        self.total_case = 0
        self.total_fail_case = 0
        self.total_online_fail_case = 0
        
        try:
            # 执行测试用例
            self._execute_test_cases(run_online_case_only)
            
            # 处理测试结果
            self._process_test_results()
            
            # 保存失败用例到数据库
            try:
                mysql_operation.batch_insert_fail_cases(self.total_fail_case_info)
            except Exception as e:
                print(f"保存失败用例到数据库时出错: {str(e)}")
            
            # 生成并发送消息
            summary_message = self._generate_summary_message(run_online_case_only)
            online_message, offline_message = self._generate_fail_case_messages()
            
            # 发送消息
            self.send_message(summary_message + online_message, send_online_message, 'wechat')
            self.send_message(offline_message, False)
            
        except Exception as e:
            error_message = f"执行测试过程中发生错误: {str(e)}"
            print(error_message)
            self.send_message(error_message, send_online_message, 'wechat')



if __name__ == "__main__":
    apifox_test = apifox_auto_test()
    
    # 执行所有用例（多线程并行执行）
    apifox_test.total_test()
    
    # 只执行线上用例
    # apifox_test.total_test(run_online_case_only=True)
    
    # 执行特定项目的用例
    # apifox_test.total_test(project_keywords=["用户管理", "订单系统"])
    
    # 执行特定项目的线上用例
    # apifox_test.total_test(run_online_case_only=True, project_keywords=["支付系统"])
